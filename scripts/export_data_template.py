"""
Export `Data Template_vCleaned.xlsx` (metadata catalog) to JSON for the CXO dashboard.

Sheets used:
  - Metadata - Data Set
  - Metadata - Column  (sampled: up to MAX_COLS_PER_DATASET columns per data set)

Output:
  - ../data/scb_data_catalog.json
"""
from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "Data Template_vCleaned.xlsx"
OUT = ROOT / "data" / "scb_data_catalog.json"

MAX_COLS_PER_DATASET = 12
MAX_DATASETS_WITH_COLUMNS = 35
MAX_DEF_LEN = 220
MAX_SAMPLE_LEN = 160
MAX_DATASETS_IN_ONTOLOGY_NODES = 35


def _norm(s: object) -> str:
    if s is None:
        return ""
    return str(s).strip()


def load_workbook_safe(xlsx: Path):
    if not xlsx.exists():
        raise FileNotFoundError(str(xlsx))
    return openpyxl.load_workbook(xlsx, data_only=True)


def export(xlsx: Path | None = None) -> dict:
    xlsx_path = xlsx if xlsx is not None else XLSX
    wb = load_workbook_safe(xlsx_path)
    if "Metadata - Data Set" not in wb.sheetnames or "Metadata - Column" not in wb.sheetnames:
        raise RuntimeError(f"Unexpected sheets: {wb.sheetnames}")

    ws_ds = wb["Metadata - Data Set"]
    ds_rows = list(ws_ds.iter_rows(values_only=True))
    ds_header = [_norm(x) for x in ds_rows[0]]
    ds_records: list[dict] = []
    for row in ds_rows[1:]:
        if not any(row):
            continue
        rec = {ds_header[i]: row[i] for i in range(min(len(ds_header), len(row)))}
        name = _norm(rec.get("Data Set Name"))
        if not name:
            continue
        ds_records.append(
            {
                "dataSetName": name,
                "dataSetFullName": _norm(rec.get("Data Set Full Name")),
                "description": _norm(rec.get("Data Set  Description")),
                "updateFrequency": _norm(rec.get("Data Set Update Frequency")),
                "subjectArea": _norm(rec.get("Data Subject Area Name")),
            }
        )

    ws_col = wb["Metadata - Column"]
    col_rows = list(ws_col.iter_rows(values_only=True))
    col_header = [_norm(x) for x in col_rows[0]]
    idx = {h: i for i, h in enumerate(col_header)}

    required_keys = [
        "Data Set Name",
        "Column Name",
        "Type",
        "Business Attribute Name",
        "Business Definition",
        "Sample Data",
        "Schema Name",
        "System / Catalog Name",
    ]
    for k in required_keys:
        if k not in idx:
            raise RuntimeError(f"Missing expected column '{k}' in Metadata - Column")

    cols_by_ds: dict[str, list[dict]] = defaultdict(list)
    allowed_ordered: list[str] = []
    allowed_set: set[str] = set()

    for row in col_rows[1:]:
        if not any(row):
            continue
        def get(k: str):
            i = idx[k]
            return row[i] if i < len(row) else None

        ds_name = _norm(get("Data Set Name"))
        if not ds_name:
            continue
        if ds_name not in allowed_set:
            if len(allowed_ordered) >= MAX_DATASETS_WITH_COLUMNS:
                continue
            allowed_ordered.append(ds_name)
            allowed_set.add(ds_name)
        if len(cols_by_ds[ds_name]) >= MAX_COLS_PER_DATASET:
            continue

        sample = get("Sample Data")
        if isinstance(sample, str) and len(sample) > MAX_SAMPLE_LEN:
            sample = sample[:MAX_SAMPLE_LEN] + "…"

        cols_by_ds[ds_name].append(
            {
                "columnName": _norm(get("Column Name")),
                "type": _norm(get("Type")),
                "businessAttributeName": _norm(get("Business Attribute Name")),
                "businessDefinition": _norm(get("Business Definition"))[:MAX_DEF_LEN],
                "sampleData": sample,
                "schemaName": _norm(get("Schema Name")),
                "systemCatalogName": _norm(get("System / Catalog Name")),
            }
        )

    ds_meta = {d["dataSetName"]: d for d in ds_records}

    # Ontology: lightweight graph for UI (only datasets we exported columns for)
    ontology_nodes = []
    ontology_edges = []
    for ds_name in allowed_ordered[:MAX_DATASETS_IN_ONTOLOGY_NODES]:
        ds = ds_meta.get(ds_name, {"dataSetName": ds_name, "dataSetFullName": "", "subjectArea": "", "updateFrequency": ""})
        ontology_nodes.append(
            {
                "id": f"dataset:{ds['dataSetName']}",
                "type": "DataSet",
                "label": ds["dataSetName"],
                "subjectArea": ds.get("subjectArea", ""),
                "updateFrequency": ds.get("updateFrequency", ""),
            }
        )
        for c in cols_by_ds.get(ds_name, [])[: min(8, MAX_COLS_PER_DATASET)]:
            ontology_nodes.append(
                {
                    "id": f"column:{ds_name}:{c['columnName']}",
                    "type": "Column",
                    "label": c["columnName"],
                    "parentDataSet": ds_name,
                    "businessAttributeName": c["businessAttributeName"],
                }
            )

    for ds_name in allowed_ordered[:MAX_DATASETS_IN_ONTOLOGY_NODES]:
        d_id = f"dataset:{ds_name}"
        for c in cols_by_ds.get(ds_name, [])[: min(8, MAX_COLS_PER_DATASET)]:
            c_id = f"column:{ds_name}:{c['columnName']}"
            ontology_edges.append({"from": d_id, "to": c_id, "rel": "hasColumn"})

    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceFile": str(xlsx_path.name),
        "dataSets": ds_records,
        "columnsByDataSet": {k: cols_by_ds[k] for k in allowed_ordered if k in cols_by_ds},
        "columnExport": {
            "maxDataSets": MAX_DATASETS_WITH_COLUMNS,
            "maxColsPerDataSet": MAX_COLS_PER_DATASET,
            "dataSetsIncluded": allowed_ordered,
        },
        "ontology": {"nodes": ontology_nodes, "edges": ontology_edges},
        "tables": [
            {
                "name": "metadata_data_set",
                "description": "Catalog of data sets (from Excel sheet Metadata - Data Set)",
                "rowCount": len(ds_records),
            },
            {
                "name": "metadata_column",
                "description": "Column-level metadata (from Excel sheet Metadata - Column; truncated per data set in JSON)",
                "rowCount": sum(len(v) for v in cols_by_ds.values()),
            },
        ],
    }
    wb.close()
    return payload


def main():
    p = argparse.ArgumentParser(description="Export Data Template xlsx to data/scb_data_catalog.json")
    p.add_argument(
        "--xlsx",
        type=Path,
        default=None,
        help=f"Path to Excel template (default: {XLSX})",
    )
    args = p.parse_args()
    xlsx = args.xlsx
    OUT.parent.mkdir(parents=True, exist_ok=True)
    data = export(xlsx)
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT} ({OUT.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
