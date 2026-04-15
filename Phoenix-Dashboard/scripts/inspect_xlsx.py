import openpyxl

path = r"C:\Users\Apurba Mukherjee\OneDrive - McKinsey & Company\Desktop\AM\SCBX\Data Template_vCleaned.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print("sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print("\n===", name, "max_row", ws.max_row, "max_col", ws.max_column)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        print(row)
        if i >= 4:
            break
wb.close()
